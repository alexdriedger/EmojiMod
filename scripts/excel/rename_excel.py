from openpyxl import load_workbook
import json
wb = load_workbook('Slay the Spire Emoji Translations.xlsx')
ws = wb.active
col_a = ws['A']
col_b = ws['B']
cards = dict()
with open('cards.json', 'r') as json_file:
  cards = json.load(json_file)
json_dict = dict()
for row in range(1, 371):
  card_name = col_a[row].value
  emoji_name = col_b[row].value
  if card_name in cards:
    mini_dict = dict()
    mini_dict["NAME"] = emoji_name
    json_dict[card_name] = mini_dict
with open('out_cards.json', 'w') as out_json_file:
  json.dump(json_dict, out_json_file, indent=2)
